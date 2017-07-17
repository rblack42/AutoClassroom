from openpyxl import Workbook
wb = Workbook()
ws1 = wb.active
ws1.title = 'Summary'
ws2 = wb.create_sheet(title='Labs')
ws3 = wb.create_sheet(title='Exam1')
ws4 = wb.create_sheet(title='Exam2')

wb.save(filename = 'cosc2325-001-SU17.xlsx')


